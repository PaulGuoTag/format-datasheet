import streamlit as st
from openpyxl import load_workbook
import io
import zipfile

# 页面配置
st.set_page_config(page_title="Excel 批量清洗工具", layout="centered")

def process_excel(file_content):
    """复刻并强化 VBA 处理逻辑"""
    # 载入文件流
    wb = load_workbook(io.BytesIO(file_content))
    
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # 1. 预处理：将 Excel 常见的特殊空格 (\xa0) 转为标准空格
                    val = cell.value.replace('\xa0', ' ')
                    
                    # 2. 步骤 A: 替换 [*] 为 /
                    # 覆盖带空格和不带空格的两种情况
                    val = val.replace(" [*] ", "/")
                    val = val.replace("[*]", "/")
                    
                    # 3. 步骤 B: 循环清理开头（针对任意空格+斜杠）
                    # 类似于 VBA 中的 LTrim + Left 检查，但更彻底
                    while True:
                        temp = val.lstrip() # 去掉左侧所有空格
                        if temp.startswith("/"):
                            val = temp[1:]  # 去掉开头的斜杠，继续下一轮检查
                        else:
                            val = temp      # 已经没有空格或斜杠了，保持现状
                            break           # 退出循环
                    
                    cell.value = val
    
    # 将处理后的文件保存到内存
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# --- Streamlit UI ---
st.title("🚀 Excel 数据清洗工具")
st.info("替换说明：将所有的 `[*]` 替换为 `/`，并清除单元格开头多余的空格与斜杠。")

# 多文件上传
uploaded_files = st.file_uploader("上传 Excel 文件 (支持拖拽多选)", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:
    processed_files = {} # 存储结果：{文件名: 二进制数据}
    
    # 进度显示
    progress_bar = st.progress(0)
    for index, uploaded_file in enumerate(uploaded_files):
        with st.spinner(f"正在处理: {uploaded_file.name}"):
            file_bytes = uploaded_file.read()
            # 执行核心处理逻辑
            output_data = process_excel(file_bytes)
            processed_files[f"processed_{uploaded_file.name}"] = output_data
            
        # 更新进度条
        progress_bar.progress((index + 1) / len(uploaded_files))

    st.success("✨ 所有文件处理完成！")

    # 下载逻辑
    if len(processed_files) == 1:
        # 单文件下载
        file_name, data = list(processed_files.items())[0]
        st.download_button(
            label="💾 下载处理后的 Excel",
            data=data,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    else:
        # 多文件打包成 ZIP 下载
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED) as zf:
            for name, data in processed_files.items():
                zf.writestr(name, data)
        
        st.download_button(
            label=f"📦 一键下载所有文件 ({len(processed_files)}个) 的压缩包",
            data=zip_buffer.getvalue(),
            file_name="batch_processed_files.zip",
            mime="application/zip",
            type="primary"
        )

st.divider()
st.caption("使用说明：将此代码和 requirements.txt 上传至 GitHub 仓库，连接 Streamlit Cloud 即可使用。")